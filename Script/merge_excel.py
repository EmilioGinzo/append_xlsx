# este script anexa los archivos te tipo .xlsx que esten en la carpeta ..\Input
# el archivo ya unido se encontrara en la carpeta ..\Output con el nombre de "Planilla de Horas.xlsx"
# Si bien es mas facil unir varios archivos xlsx utilizando .append de pandas
# decidi hacerlo de la siguiente forma para que el archivo final tenga un formato similar 
# a los archivos de input

import os
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment,Protection

IN_PATH = '..\Input\\'

#da estilo a la hoja que le pasemos como parametro
def set_styles(ws: Worksheet) -> None:
    #aqui podemos cabiar el estilo del excel final
    fill = PatternFill(fill_type = 'solid',start_color = 'F2F2F2F2', end_color = 'F2F2F2F2')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center')

    #le damos el estilo correspondiente
    for row in ws:
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            cell.fill = fill

#recibe la direccion y el nombre del archivo xlsx 
#crea y retorna un array de arrays que contiene el archivo xlsx 
# ejemplo: array = [['Dana Kennedy', '9', 'Octubre', '2017', '2', '1019', 'FACTURABLE'], ['Cameron Weiss', '23', 'Octubre', '2017', '1', '1021', 'NO FACTURABLE'], ...]
def append_array(file: str) -> list:
    book = load_workbook(IN_PATH + file)
    array = []
    for row in book.active.rows:
            array2 = []
            for cell in row:
                array2.append(cell.value)
            array.append(array2)
    return array

#recibe la hoja activa en la cual se quieren agregar los datos y los archivos a agregar
#agrega los datos de los archivos files_xlsx a la hoja ws
def append_xlsx(ws: Worksheet, files_xlsx: list) -> None:
    for file in files_xlsx[1:]:
        array = append_array(file)
        for row in array[1:]:
            ws.append(row)

if __name__ == '__main__':
    if not os.path.exists('..\Output'):
        os.makedirs('..\Output')
    if not os.path.exists('..\Input'):
        os.makedirs('..\Input')
    with pd.ExcelWriter('..\Output\Planilla de Horas.xlsx', engine='openpyxl') as writer:
        files = os.listdir(IN_PATH)
        files_xlsx = [f for f in files if f[-4:] == 'xlsx']
        df_result = pd.DataFrame()
        writer.book = load_workbook(IN_PATH + files_xlsx[0])
        ws = writer.book.active

        append_xlsx(ws, files_xlsx)
        set_styles(ws)

        df_result.to_excel(writer, sheet_name = 'Horas')